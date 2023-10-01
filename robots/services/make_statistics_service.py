from robots.models import Robot
from django.utils import timezone
from django.db.models import Count
from openpyxl import Workbook


class MakeReportService:
    def _get_data_for_report(self) -> list[dict]:
        self.to = timezone.now()
        self.fr = self.to - timezone.timedelta(days=7)
        robots_for_week = Robot.objects.filter(created__range=[self.fr, self.to])

        queryset = robots_for_week.values('serial', 'model', 'version').order_by('serial')
        counted_queryset = queryset.annotate(amount=Count('serial'))

        return list(counted_queryset)

    def _get_path(self) -> str:
        return f"Report from {self.fr.date()} to {self.to.date()}.xlsx"

    def make_report(self) -> str:
        wb = Workbook()
        data_for_report = self._get_data_for_report()
        for data in data_for_report:
            model = data['model']
            version = data['version']
            amount = data['amount']
            robot_data = {'A': model, 'B': version, 'C': amount}

            if model not in wb.sheetnames:
                ws = wb.create_sheet(model)
                ws['A1'] = 'Модель'
                ws['B1'] = 'Версия'
                ws['C1'] = 'Количество за неделю'
                ws.append(robot_data)
            else:
                ws = wb[model]
                ws.append(robot_data)

        path = f"robots/services/{self._get_path()}"
        wb.remove(wb['Sheet'])
        wb.save(path)
        wb.close()

        return path
